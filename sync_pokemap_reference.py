#!/usr/bin/env python3
# sync_pokemap_reference.py  — v2.1
# Columnas nuevas: NivelBase, CategoriaGO, EtiquetasGO, esBebe, Region
# Export: CSV + Excel (pokeMAP.xlsx)
import csv
import hashlib
import json
import os
import shutil
import subprocess
from datetime import datetime, timezone

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


SOURCES = {
    "pokemon_stats": "https://pogoapi.net/api/v1/pokemon_stats.json",
    "pokemon_types": "https://pogoapi.net/api/v1/pokemon_types.json",
    "released_pokemon": "https://pogoapi.net/api/v1/released_pokemon.json",
    "pokemon_evolutions": "https://pogoapi.net/api/v1/pokemon_evolutions.json",
    "mega_pokemon": "https://pogoapi.net/api/v1/mega_pokemon.json",
    "shadow_pokemon": "https://pogoapi.net/api/v1/shadow_pokemon.json",
    "raid_bosses": "https://pogoapi.net/api/v1/raid_bosses.json",
    "pokemon_species_csv": "https://raw.githubusercontent.com/veekun/pokedex/master/pokedex/data/csv/pokemon_species.csv",
}

# IDs de Ultra Bestias (estables entre generaciones)
ULTRA_BEAST_IDS = {793, 794, 795, 796, 797, 798, 799, 803, 804, 805, 806}

# Regionales en Pokémon GO: {pokemon_id: "Región"}
# Fuente: conocimiento consolidado de la comunidad GO (aproximado, cambia con eventos)
REGIONAL_MAP = {
    # Gen 1
    83:  "Asia Oriental",
    115: "Australia / Nueva Zelanda",
    122: "Europa",
    128: "Norteamérica",
    # Gen 2
    214: "Sudamérica / Sur de EEUU",
    222: "Regiones tropicales",
    313: "Europa / Asia / Pacífico",
    314: "América / África",
    # Gen 3
    324: "Sur de Asia",
    335: "Europa / Asia / Pacífico",
    336: "América / África",
    337: "América / África",
    338: "Europa / Asia / Pacífico",
    357: "África / Sur de Europa",
    369: "Nueva Zelanda / Oceanía",
    # Gen 4
    417: "Canadá / Norte de EEUU",
    422: "Global (formas por región)",  # Shellos
    423: "Global (formas por región)",  # Gastrodon
    439: "Europa",
    441: "Hemisferio Sur",
    455: "Sureste de EEUU",
    # Legendarios regionales (raids, no captura libre)
    480: "Asia-Pacífico",   # Uxie
    481: "Europa/África",   # Mesprit
    482: "América",         # Azelf
    # Gen 5
    538: "Europa / Asia / Pacífico",
    539: "América / África",
    556: "América",
    561: "Egipto / Grecia",
    626: "Norteamérica",
    631: "América / África / Oriente Medio",
    632: "Europa / Asia / Pacífico",
    # Gen 6
    707: "Asia",            # Klefki
    # Gen 7
    741: "Global (formas por región)",  # Oricorio
    # Gen 8
    865: "Gallar/Reino Unido",  # Sirfetch'd (event, no estrictamente regional)
    875: "Global",              # Eiscue
    # Gen 9 (expandir según disponibilidad GO)
}

# Tabla de efectividad: {tipo_defensor: [tipos_super_efectivos]}
SUPER_EFFECTIVE = {
    "normal": ["fighting"],
    "fire": ["water", "ground", "rock"],
    "water": ["electric", "grass"],
    "electric": ["ground"],
    "grass": ["fire", "ice", "poison", "flying", "bug"],
    "ice": ["fire", "fighting", "rock", "steel"],
    "fighting": ["flying", "psychic", "fairy"],
    "poison": ["ground", "psychic"],
    "ground": ["water", "grass", "ice"],
    "flying": ["electric", "ice", "rock"],
    "psychic": ["bug", "ghost", "dark"],
    "bug": ["fire", "flying", "rock"],
    "rock": ["water", "grass", "fighting", "ground", "steel"],
    "ghost": ["ghost", "dark"],
    "dragon": ["ice", "dragon", "fairy"],
    "dark": ["fighting", "bug", "fairy"],
    "steel": ["fire", "water", "ground"],
    "fairy": ["poison", "steel"],
}

# Inversa: {tipo_atacante: [tipos_defendidos_super_efectivamente]}
DEALS_SUPER_TO = {}
for defender, attackers in SUPER_EFFECTIVE.items():
    for attacker in attackers:
        if attacker not in DEALS_SUPER_TO:
            DEALS_SUPER_TO[attacker] = []
        DEALS_SUPER_TO[attacker].append(defender)

TYPE_ES = {
    "Normal": "normal",
    "Fire": "fuego",
    "Water": "agua",
    "Electric": "electrico",
    "Grass": "planta",
    "Ice": "hielo",
    "Fighting": "lucha",
    "Poison": "veneno",
    "Ground": "tierra",
    "Flying": "volador",
    "Psychic": "psiquico",
    "Bug": "bicho",
    "Rock": "roca",
    "Ghost": "fantasma",
    "Dragon": "dragon",
    "Dark": "siniestro",
    "Steel": "acero",
    "Fairy": "hada",
}


def ensure_dirs():
    os.makedirs("data/raw", exist_ok=True)
    os.makedirs("data/metadata", exist_ok=True)
    os.makedirs("data/reports", exist_ok=True)


def run_curl(url, output_path):
    subprocess.run(["curl", "-fsSL", url, "-o", output_path], check=True)


def sha256_file(path):
    digest = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_species_data(path):
    """Devuelve dict {pid: {nivel_base, es_bebe, evolves_from_id}} desde species CSV."""
    data = {}
    with open(path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                pid = int(row["id"])
            except Exception:
                continue
            evolves_from_raw = row.get("evolves_from_species_id", "")
            evolves_from_id = int(evolves_from_raw) if evolves_from_raw else None
            is_legendary = row.get("is_legendary", "0") == "1"
            is_mythical = row.get("is_mythical", "0") == "1"
            is_baby = row.get("is_baby", "0") == "1"
            if is_mythical:
                nivel = "singular"
            elif is_legendary:
                nivel = "legendario"
            else:
                nivel = "normal"
            data[pid] = {
                "nivel_base": nivel,
                "es_bebe": is_baby,
                "evolves_from_id": evolves_from_id,
            }
    return data


def pick_preferred_form(rows):
    by_id = {}
    for row in rows:
        pid = int(row["pokemon_id"])
        current = by_id.get(pid)
        if current is None or row.get("form") == "Normal":
            by_id[pid] = row
    return by_id


def build_mega_sets(mega_data):
    """Devuelve (set_ids_con_mega, set_ids_con_primigenio)."""
    mega_ids = set()
    primal_ids = set()
    for entry in mega_data:
        pid = int(entry["pokemon_id"])
        mega_ids.add(pid)
        if "primal" in entry.get("mega_name", "").lower():
            primal_ids.add(pid)
    return mega_ids, primal_ids


def build_shadow_set(shadow_data):
    """Devuelve set de IDs con forma sombra disponible en GO."""
    return {int(k) for k in shadow_data.keys()}


def build_raid_sets(raid_data):
    """Devuelve {pid: tier_minimo} para raid bosses actuales."""
    raid_ids = {}
    if not isinstance(raid_data, dict) or "current" not in raid_data:
        return raid_ids
    for tier_str, bosses in raid_data.get("current", {}).items():
        try:
            tier = int(tier_str)
        except ValueError:
            continue
        for boss in bosses:
            pid = int(boss.get("id", 0))
            if pid > 0:
                if pid not in raid_ids or tier < raid_ids[pid]:
                    raid_ids[pid] = tier
    return raid_ids


def build_children_map(species_data, released_ids):
    """Crea mapa {parent_id: [child_ids]} limitado al pool de Pokémon released en GO."""
    children = {}
    for pid, sp in species_data.items():
        parent = sp.get("evolves_from_id")
        if parent is None:
            continue
        if pid not in released_ids or parent not in released_ids:
            continue
        if parent not in children:
            children[parent] = []
        children[parent].append(pid)
    for parent in children:
        children[parent] = sorted(children[parent])
    return children


def build_evolution_graph(evo_data, released_ids):
    """Crea grafo {from_id: [{to_id, candy, item}]} limitado a Pokémon released en GO."""
    edge_map = {}
    for entry in evo_data:
        from_id = int(entry.get("pokemon_id", 0))
        if from_id == 0 or from_id not in released_ids:
            continue
        for evo in entry.get("evolutions", []):
            to_id = int(evo.get("pokemon_id", 0))
            if to_id == 0 or to_id not in released_ids:
                continue
            candy = int(evo.get("candy_required", 0) or 0)
            item = (evo.get("item_required") or "").strip()
            key = (from_id, to_id)
            # Si hay duplicados por formas, quedarse con el menor coste de caramelos.
            if key not in edge_map or candy < edge_map[key]["candy"]:
                edge_map[key] = {"to_id": to_id, "candy": candy, "item": item}

    graph = {}
    for (from_id, _), edge in edge_map.items():
        if from_id not in graph:
            graph[from_id] = []
        graph[from_id].append(edge)
    for from_id in graph:
        graph[from_id] = sorted(graph[from_id], key=lambda e: (e["to_id"], e["candy"]))
    return graph


def get_descendants(pid, children_map):
    """Devuelve todos los descendientes evolutivos alcanzables desde pid."""
    out = []
    stack = list(children_map.get(pid, []))
    seen = set()
    while stack:
        current = stack.pop(0)
        if current in seen:
            continue
        seen.add(current)
        out.append(current)
        stack.extend(children_map.get(current, []))
    return out


def get_descendant_paths(pid, evo_graph):
    """Devuelve {target_id: {candy, items, path_ids}} con el camino más barato en caramelos."""
    best = {}
    queue = []
    for edge in evo_graph.get(pid, []):
        queue.append((edge["to_id"], edge["candy"], [edge["item"]] if edge["item"] else [], [pid, edge["to_id"]]))

    while queue:
        node, candy_sum, items, path_ids = queue.pop(0)
        prev = best.get(node)
        if prev is not None and candy_sum >= prev["candy"]:
            continue
        best[node] = {"candy": candy_sum, "items": items, "path_ids": path_ids}

        for edge in evo_graph.get(node, []):
            next_id = edge["to_id"]
            if next_id in path_ids:
                continue
            next_items = list(items)
            if edge["item"]:
                next_items.append(edge["item"])
            queue.append((next_id, candy_sum + edge["candy"], next_items, path_ids + [next_id]))

    return best


def pokemon_power_index(row):
    """Índice único para comparar potencial (pondera PvE y PvP)."""
    return int(row["ScorePvE"]) + int(row["ScorePvP_GL"]) + int(row["ScorePvP_UL"])


def score_pve(attack, defense, stamina, tipos_json):
    """Score PvE (1-100) basado en stats."""
    # Normalizar a 0-100 (max realistic stats ~300 attack, ~250 def, ~300 stamina)
    atk_score = min(100, int((attack / 300.0) * 100))
    def_score = min(100, int((defense / 250.0) * 100))
    sta_score = min(100, int((stamina / 300.0) * 100))
    # PvE favorece ataque fuerte
    return int((atk_score * 0.5 + def_score * 0.2 + sta_score * 0.3))


def score_pvp_gl(attack, defense, stamina):
    """Score PvP Gran Liga (1-100). GL favorece alto CP pero stats balanceados.
    CP máx en GL: 1500. Preferencia: defensa + stamina > ataque puro."""
    # Calcular CP estimado (fórmula GO simplificada)
    cp = ((attack + 15) * (defense + 15) * (stamina + 15)) // 10000
    cp_score = min(100, max(50, int((cp / 1500.0) * 100)))
    # Balance: defensa + stamina son críticos
    def_sta = (defense + stamina) * 0.5
    balance_score = min(100, int((def_sta / 300.0) * 100))
    return int((cp_score * 0.4 + balance_score * 0.6))


def score_pvp_ul(attack, defense, stamina):
    """Score PvP Ultra Liga (1-100). UL permite CP hasta 2500.
    Preferencia: ligeramente más ataque que GL, pero balance sigue siendo clave."""
    cp = ((attack + 15) * (defense + 15) * (stamina + 15)) // 10000
    cp_score = min(100, max(50, int((cp / 2500.0) * 100)))
    atk_score = min(100, int((attack / 300.0) * 100))
    balance_score = min(100, int(((defense + stamina) * 0.5 / 300.0) * 100))
    return int((cp_score * 0.3 + atk_score * 0.4 + balance_score * 0.3))


def determine_main_use(score_pve, score_pvp_gl, score_pvp_ul, raid_tier, es_raid):
    """Determina uso principal basado en scores."""
    if es_raid and raid_tier == 1:
        return "raid"
    if score_pve >= 75:
        return "pve"
    if score_pvp_gl >= 70:
        return "pvp_gl"
    if score_pvp_ul >= 70:
        return "pvp_ul"
    return "coleccionable"


def get_counters(tipos_str):
    """Devuelve tipos contra los que es fuerte (como string separado por /)."""
    if not tipos_str or tipos_str == "desconocido":
        return ""
    tipos = tipos_str.split("/")
    strong_against = set()
    for tipo in tipos:
        tipo_en = None
        for en, es in TYPE_ES.items():
            if es == tipo.lower():
                tipo_en = en.lower()
                break
        if tipo_en and tipo_en in DEALS_SUPER_TO:
            strong_against.update(DEALS_SUPER_TO[tipo_en])
    if not strong_against:
        return ""
    # Traducir de vuelta a español
    result = []
    for t in sorted(strong_against):
        es = TYPE_ES.get(t.capitalize(), t)
        result.append(es)
    return "/".join(result[:3])  # Top 3 solo


def build_rows(stats, types, released, evolutions_data, mega_data, shadow_data, raid_data, species_data):
    stats_by_id = pick_preferred_form(stats)
    types_by_id = pick_preferred_form(types)
    mega_ids, primal_ids = build_mega_sets(mega_data)
    shadow_ids = build_shadow_set(shadow_data)
    raid_ids = build_raid_sets(raid_data)

    fecha = datetime.now().strftime("%Y-%m-%d")
    released_ids = {int(k) for k in released.keys()}
    children_map = build_children_map(species_data, released_ids)
    evo_graph = build_evolution_graph(evolutions_data, released_ids)
    base_rows = []

    for pid_str, released_info in released.items():
        pid = int(pid_str)
        st = stats_by_id.get(pid)
        tp = types_by_id.get(pid)
        if st is None:
            continue

        sp = species_data.get(pid, {})
        nivel_base = sp.get("nivel_base", "desconocido")
        es_bebe = sp.get("es_bebe", False)

        # CategoriaGO: clasificación jerárquica GO-específica
        if pid in ULTRA_BEAST_IDS:
            categoria_go = "ultraente"
        elif nivel_base == "singular":
            categoria_go = "singular"
        elif nivel_base == "legendario":
            categoria_go = "legendario"
        elif es_bebe:
            categoria_go = "bebe"
        else:
            categoria_go = "normal"

        # EtiquetasGO: multi-etiqueta separada por comas
        etiquetas = []
        if es_bebe:
            etiquetas.append("bebe")
        if pid in mega_ids:
            etiquetas.append("mega")
        if pid in primal_ids:
            etiquetas.append("primigenio")
        if pid in shadow_ids:
            etiquetas.append("sombra")
        if pid in REGIONAL_MAP:
            etiquetas.append("regional")

        # Región (solo para regionales)
        region = REGIONAL_MAP.get(pid, "")

        tipo_raw = tp.get("type", []) if tp else []
        tipo = "/".join(TYPE_ES.get(t, t.lower()) for t in tipo_raw) if tipo_raw else "desconocido"
        nombre = st.get("pokemon_name") or released_info.get("name") or f"Pokemon {pid}"

        # Stats
        attack = int(st.get("base_attack", 0))
        defense = int(st.get("base_defense", 0))
        stamina = int(st.get("base_stamina", 0))

        # Scores
        pve_score = score_pve(attack, defense, stamina, tipo)
        pvp_gl_score = score_pvp_gl(attack, defense, stamina)
        pvp_ul_score = score_pvp_ul(attack, defense, stamina)

        # Raid
        es_raid = "sí" if pid in raid_ids else "no"
        raid_tier = raid_ids.get(pid, 0)

        # Uso principal
        uso_principal = determine_main_use(pve_score, pvp_gl_score, pvp_ul_score, raid_tier, es_raid == "sí")

        # Contadores
        contadores = get_counters(tipo)

        base_rows.append(
            {
                "id": pid,
                "Nombre": nombre,
                "Tipo": tipo,
                "NivelBase": nivel_base,
                "CategoriaGO": categoria_go,
                "EtiquetasGO": ",".join(etiquetas),
                "esBebe": "sí" if es_bebe else "no",
                "Region": region,
                "Ataque Base": attack,
                "Defensa Base": defense,
                "Stamina Base": stamina,
                "ScorePvE": pve_score,
                "ScorePvP_GL": pvp_gl_score,
                "ScorePvP_UL": pvp_ul_score,
                "UsoPrincipal": uso_principal,
                "Contadores": contadores,
                "EsRaid": es_raid,
                "fechaActualizacion": fecha,
            }
        )

    # Segunda pasada: potencial por evolución
    rows_by_id = {row["id"]: row for row in base_rows}
    for row in base_rows:
        pid = row["id"]
        direct_evos = [e["to_id"] for e in evo_graph.get(pid, [])]
        paths = get_descendant_paths(pid, evo_graph)
        all_desc = sorted(paths.keys()) if paths else get_descendants(pid, children_map)

        row["EvolucionaA"] = "/".join(rows_by_id[x]["Nombre"] for x in direct_evos if x in rows_by_id)

        if not all_desc:
            row["EvolucionRecomendada"] = ""
            row["UsoEvolucionRecomendada"] = ""
            row["ScoreEvolucionRecomendada"] = ""
            row["CosteCaramelosEvolucionRecomendada"] = ""
            row["ObjetosEvolucionRecomendada"] = ""
            continue

        candidates = [pid] + [x for x in all_desc if x in rows_by_id]
        best_id = max(candidates, key=lambda x: pokemon_power_index(rows_by_id[x]))
        current_score = pokemon_power_index(rows_by_id[pid])
        best_score = pokemon_power_index(rows_by_id[best_id])

        # Solo recomendar evolución si mejora realmente el potencial
        if best_id != pid and best_score > current_score:
            row["EvolucionRecomendada"] = rows_by_id[best_id]["Nombre"]
            row["UsoEvolucionRecomendada"] = rows_by_id[best_id]["UsoPrincipal"]
            row["ScoreEvolucionRecomendada"] = best_score
            best_path = paths.get(best_id, {"candy": "", "items": []})
            row["CosteCaramelosEvolucionRecomendada"] = best_path.get("candy", "")
            items = sorted(set(i for i in best_path.get("items", []) if i))
            row["ObjetosEvolucionRecomendada"] = "/".join(items)
        else:
            row["EvolucionRecomendada"] = ""
            row["UsoEvolucionRecomendada"] = ""
            row["ScoreEvolucionRecomendada"] = ""
            row["CosteCaramelosEvolucionRecomendada"] = ""
            row["ObjetosEvolucionRecomendada"] = ""

    rows = base_rows
    rows.sort(key=lambda r: r["id"])
    return rows


CSV_FIELDS = [
    "Nombre",
    "Tipo",
    "NivelBase",
    "CategoriaGO",
    "EtiquetasGO",
    "esBebe",
    "Region",
    "Ataque Base",
    "Defensa Base",
    "Stamina Base",
    "ScorePvE",
    "ScorePvP_GL",
    "ScorePvP_UL",
    "UsoPrincipal",
    "Contadores",
    "EsRaid",
    "EvolucionaA",
    "EvolucionRecomendada",
    "UsoEvolucionRecomendada",
    "ScoreEvolucionRecomendada",
    "CosteCaramelosEvolucionRecomendada",
    "ObjetosEvolucionRecomendada",
    "fechaActualizacion",
]

SNAPSHOT_FIELDS = ["Nombre", "Tipo", "NivelBase", "CategoriaGO", "EtiquetasGO",
                   "esBebe", "Region", "Ataque Base", "Defensa Base", "Stamina Base",
                   "ScorePvE", "ScorePvP_GL", "ScorePvP_UL", "UsoPrincipal", "Contadores", "EsRaid",
                   "EvolucionaA", "EvolucionRecomendada", "UsoEvolucionRecomendada", "ScoreEvolucionRecomendada",
                   "CosteCaramelosEvolucionRecomendada", "ObjetosEvolucionRecomendada"]


def write_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row[k] for k in CSV_FIELDS})


def write_excel(rows, path):
    """Escribe el CSV a un archivo .xlsx con estilos básicos. Requiere openpyxl."""
    if not HAS_OPENPYXL:
        print(f"⚠️  openpyxl no está instalado. Skipping {path}")
        print("   Para habilitar: pip install openpyxl")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "PokeMAP"

    # Header con estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, field in enumerate(CSV_FIELDS, 1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Datos
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, field in enumerate(CSV_FIELDS, 1):
            value = row_data.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            # Color alternado para mejor legibilidad
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # Auto-ajuste de anchos
    ws.column_dimensions["A"].width = 20   # Nombre
    ws.column_dimensions["B"].width = 15   # Tipo
    ws.column_dimensions["C"].width = 12   # NivelBase
    ws.column_dimensions["D"].width = 14   # CategoriaGO
    ws.column_dimensions["E"].width = 25   # EtiquetasGO
    ws.column_dimensions["F"].width = 8    # esBebe
    ws.column_dimensions["G"].width = 18   # Region
    ws.column_dimensions["H"].width = 12   # Ataque Base
    ws.column_dimensions["I"].width = 12   # Defensa Base
    ws.column_dimensions["J"].width = 12   # Stamina Base
    ws.column_dimensions["K"].width = 10   # ScorePvE
    ws.column_dimensions["L"].width = 12   # ScorePvP_GL
    ws.column_dimensions["M"].width = 12   # ScorePvP_UL
    ws.column_dimensions["N"].width = 14   # UsoPrincipal
    ws.column_dimensions["O"].width = 25   # Contadores
    ws.column_dimensions["P"].width = 8    # EsRaid
    ws.column_dimensions["Q"].width = 18   # EvolucionaA
    ws.column_dimensions["R"].width = 22   # EvolucionRecomendada
    ws.column_dimensions["S"].width = 20   # UsoEvolucionRecomendada
    ws.column_dimensions["T"].width = 12   # ScoreEvolucionRecomendada
    ws.column_dimensions["U"].width = 14   # CosteCaramelosEvolucionRecomendada
    ws.column_dimensions["V"].width = 20   # ObjetosEvolucionRecomendada
    ws.column_dimensions["W"].width = 14   # fechaActualizacion

    # Freeze panes en header
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"✓ Excel generado: {path}")


def rows_by_id(rows):
    out = {}
    for row in rows:
        out[str(row["id"])] = {f: row[f] for f in SNAPSHOT_FIELDS}
    return out


def diff_hashes(previous, current):
    if not previous:
        return list(current.keys())
    changed = []
    for key, digest in current.items():
        if previous.get(key) != digest:
            changed.append(key)
    return sorted(changed)


def diff_rows(previous_rows, current_rows):
    prev_ids = set(previous_rows.keys())
    curr_ids = set(current_rows.keys())
    added = sorted(curr_ids - prev_ids, key=lambda x: int(x))
    removed = sorted(prev_ids - curr_ids, key=lambda x: int(x))

    changed = []
    shared = sorted(prev_ids & curr_ids, key=lambda x: int(x))
    for pid in shared:
        old = previous_rows[pid]
        new = current_rows[pid]
        field_changes = {}
        for field in SNAPSHOT_FIELDS:
            if old.get(field) != new.get(field):
                field_changes[field] = {"old": old.get(field), "new": new.get(field)}
        if field_changes:
            changed.append({"id": int(pid), "changes": field_changes})

    return added, removed, changed


def write_report(path, now_iso, source_changes, added, removed, changed):
    lines = []
    lines.append("# Delta PokeMAP")
    lines.append("")
    lines.append(f"- Fecha ejecucion UTC: {now_iso}")
    lines.append(f"- Fuentes con hash cambiado: {len(source_changes)}")
    lines.append(f"- Pokemon nuevos: {len(added)}")
    lines.append(f"- Pokemon eliminados: {len(removed)}")
    lines.append(f"- Pokemon modificados: {len(changed)}")
    lines.append("")

    if source_changes:
        lines.append("## Fuentes cambiadas")
        for src in source_changes:
            lines.append(f"- {src}")
        lines.append("")

    if added:
        lines.append("## IDs nuevos")
        lines.append("- " + ", ".join(added[:50]))
        if len(added) > 50:
            lines.append(f"- ... y {len(added) - 50} mas")
        lines.append("")

    if removed:
        lines.append("## IDs eliminados")
        lines.append("- " + ", ".join(removed[:50]))
        if len(removed) > 50:
            lines.append(f"- ... y {len(removed) - 50} mas")
        lines.append("")

    if changed:
        lines.append("## Cambios de campos (muestra)")
        for item in changed[:100]:
            lines.append(f"- ID {item['id']}: {json.dumps(item['changes'], ensure_ascii=False)}")
        if len(changed) > 100:
            lines.append(f"- ... y {len(changed) - 100} cambios mas")
        lines.append("")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def main():
    ensure_dirs()

    now = datetime.now(timezone.utc)
    run_id = now.strftime("%Y%m%d_%H%M%SZ")
    now_iso = now.isoformat()

    raw_dir = os.path.join("data/raw", run_id)
    os.makedirs(raw_dir, exist_ok=True)

    local_files = {}
    hashes = {}

    for name, url in SOURCES.items():
        ext = ".json" if name != "pokemon_species_csv" else ".csv"
        dst = os.path.join(raw_dir, f"{name}{ext}")
        run_curl(url, dst)
        local_files[name] = dst
        hashes[name] = sha256_file(dst)

    stats = load_json(local_files["pokemon_stats"])
    types = load_json(local_files["pokemon_types"])
    released = load_json(local_files["released_pokemon"])
    evolutions_data = load_json(local_files["pokemon_evolutions"])
    mega_data = load_json(local_files["mega_pokemon"])
    shadow_data = load_json(local_files["shadow_pokemon"])
    raid_data = load_json(local_files["raid_bosses"])
    species_data = load_species_data(local_files["pokemon_species_csv"])

    rows = build_rows(stats, types, released, evolutions_data, mega_data, shadow_data, raid_data, species_data)
    write_csv(rows, "pokeMAP.csv")
    os.makedirs("docs/data", exist_ok=True)
    write_csv(rows, "docs/data/pokeMAP.csv")
    write_excel(rows, "pokeMAP.xlsx")
    current_rows = rows_by_id(rows)

    latest_hashes_path = "data/metadata/latest_hashes.json"
    previous_hashes = load_json(latest_hashes_path) if os.path.exists(latest_hashes_path) else {}
    source_changes = diff_hashes(previous_hashes, hashes)

    latest_dataset_path = "data/metadata/latest_dataset.json"
    previous_dataset = load_json(latest_dataset_path) if os.path.exists(latest_dataset_path) else {}
    added, removed, changed = diff_rows(previous_dataset, current_rows)

    save_json(latest_hashes_path, hashes)
    save_json(latest_dataset_path, current_rows)

    snapshot_hashes_path = os.path.join("data/metadata", f"hashes_{run_id}.json")
    snapshot_dataset_path = os.path.join("data/metadata", f"dataset_{run_id}.json")
    save_json(snapshot_hashes_path, hashes)
    save_json(snapshot_dataset_path, current_rows)

    history_line = {
        "run_id": run_id,
        "executed_at_utc": now_iso,
        "source_hashes": hashes,
        "source_changes": source_changes,
        "rows": len(rows),
        "added": len(added),
        "removed": len(removed),
        "changed": len(changed),
    }
    with open("data/metadata/hash_history.jsonl", "a", encoding="utf-8") as f:
        f.write(json.dumps(history_line, ensure_ascii=False) + "\n")

    report_path = os.path.join("data/reports", f"delta_{run_id}.md")
    write_report(report_path, now_iso, source_changes, added, removed, changed)
    shutil.copyfile(report_path, "data/reports/last_delta.md")

    print(f"✓ CSV generado: pokeMAP.csv ({len(rows)} filas)")
    if HAS_OPENPYXL:
        print(f"✓ Excel generado: pokeMAP.xlsx")
    else:
        print(f"⚠️  Excel no generado (instala openpyxl: pip install openpyxl)")
    print(f"✓ Hashes guardados en: {latest_hashes_path}")
    print(f"✓ Reporte delta: {report_path}")


if __name__ == "__main__":
    main()
